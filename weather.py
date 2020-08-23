from urllib.request import urlopen
from urllib.parse import urlencode, unquote, quote_plus
import urllib
import json
from properties import *
from openpyxl import load_workbook
from xml.etree.ElementTree import parse
import datetime

wb = load_workbook(exel_file_name)
ws = wb.active

# 지역명 받아서  x, y 좌표 반환
def get_xy(province):
    for row in ws.rows:
        if(row[4].value == province):
            return row[5].value, row[6].value
        elif(row.count == ws.max_row):
            return Exception(str('입력한 동네는 데이터에 없습니다!'))

# 기상청 api 에 요청 보내고 받는 함수
def request_and_response(province):
    nx, ny = get_xy(province)

    now = datetime.datetime.now()

    params = '?' + urlencode({
        quote_plus("serviceKey"): serviceKey,
        quote_plus("pageNo"): "1",
        quote_plus("numOfRows"): "0", # 최대값
        quote_plus("dataType"): "JSON",         # 응답자료형식 : XML, JSON
        quote_plus("base_date"): now.strftime('%Y%m%d'),    # 발표일자 // yyyymmdd , 자정이 넘었다면 하루 전 데이터를 불러야 한다.
        quote_plus("base_time"): "0500",        # 발표시각 // HHMM, 매 시각 40분 이후 호출
        quote_plus("nx"): nx,                # 예보지점 X 좌표
        quote_plus("ny"): ny,
    })

    req = urllib.request.Request(CallbackURL + unquote(params))
    
    response_body = urlopen(req).read()
    root = json.loads(response_body)
    return root

# 정보를 받아서 출력하는 함수
def request_and_print(province):  
    print("\n" + province + "의 날씨 예보입니다.")
    # 전체 데이터 받기
    items = request_and_response(province)['response']['body']['items']['item']

    # 공백용
    print()
    print('------------------------------------')

    start_time = '0900'  # 예보 시작시간
    start_date = datetime.datetime.now().strftime('%Y%m%d')

    print()
    print(start_date[0:4] + "년" + " " + start_date[4:6] + "월" +
          " " + start_date[6:8] + '일' + " " + start_time[0:2] + '시 예보입니다')
    print()

  # 순회돌며 예보출력
    for item in items:
       # 달라질 때마다 시간 출력 하고 시작시간으로 변경함
        if(start_time != item['fcstTime']):
            print("\n--------------------------------------------------------")
            current_time = item['fcstTime']

            if(current_time == '0000'):  # 자정 넘으면 날짜 바뀜
                start_date = item['fcstDate']

            start_time = current_time
            print("\n" + start_date[0:4] + "년" + " " + start_date[4:6] + "월" +
          " " + start_date[6:8] + '일' + " " + start_time[0:2] + '시 예보입니다\n')

        print(code[item['category']] + " " + item['fcstValue'])
        
        
